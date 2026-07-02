#!/usr/bin/env python3
"""Generate an app-ready production manifest from the OSF crosswalk."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


OUTPUT_COLUMNS = [
    "audio_file",
    "audio_url",
    "target_word",
    "participant_id",
    "l1_condition",
    "pronunciation_condition",
    "stimulus_list",
    "word_number",
    "condition",
    "talker",
    "take_number",
    "spoken_form",
    "practice_note",
    "source_format",
    "counterbalance_word_number",
    "source_word_number",
    "global_speaker_id",
    "old_speaker_id",
    "proposed_speaker_id",
    "old_relative_path",
    "osf_audio_file",
    "original_pronunciation_label",
    "source_pass_number",
    "trial_number",
    "talker_label",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def speaker_lookup(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {row["proposed_speaker_id"]: row for row in rows if row.get("proposed_speaker_id")}


def normalize_word(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def read_counterbalance_word_numbers(path: Path) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Sheet1" not in workbook.sheetnames:
        raise ValueError(f"{path} must contain Sheet1 with # and Item columns.")
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    try:
        number_index = headers.index("#")
        item_index = headers.index("Item")
    except ValueError as exc:
        raise ValueError(f"{path} Sheet1 must contain # and Item columns.") from exc

    values_by_word: dict[str, set[int]] = defaultdict(set)
    for row in rows:
        if not row or row[number_index] is None or row[item_index] is None:
            continue
        word = normalize_word(str(row[item_index]))
        if not word:
            continue
        values_by_word[word].add(int(row[number_index]))

    conflicts = {word: sorted(values) for word, values in values_by_word.items() if len(values) > 1}
    if conflicts:
        examples = ", ".join(f"{word}={values}" for word, values in list(conflicts.items())[:8])
        raise ValueError(f"CounterBalance word-number conflicts: {examples}")
    return {word: next(iter(values)) for word, values in values_by_word.items()}


def build_manifest_rows(
    file_rows: list[dict[str, str]],
    speakers_by_proposed_id: dict[str, dict[str, str]],
    counterbalance_word_numbers: dict[str, int],
    path_mode: str,
    audio_base_url: str,
) -> list[dict[str, str]]:
    manifest_rows: list[dict[str, str]] = []
    for row in file_rows:
        if row.get("stimulus_set") != "main":
            continue
        if row.get("include_in_two_condition_design") != "yes":
            continue
        pronunciation = row.get("pronunciation_condition", "")
        l1 = row.get("l1_condition", "")
        if l1 not in {"ENG", "JPN", "CHN"}:
            continue
        if l1 == "ENG" and pronunciation != "natural":
            continue
        if l1 in {"JPN", "CHN"} and pronunciation not in {"natural", "accented"}:
            continue

        osf_audio_file = row["new_relative_path"]
        current_audio_file = row["old_relative_path"]
        audio_file = osf_audio_file if path_mode == "osf" else current_audio_file
        speaker = speakers_by_proposed_id.get(row.get("proposed_speaker_id", ""), {})
        global_speaker_id = speaker.get("global_speaker_id", "")
        audio_url = f"{audio_base_url.rstrip('/')}/{audio_file}" if audio_base_url else ""
        target_word = normalize_word(row["target_word"])
        if target_word not in counterbalance_word_numbers:
            raise ValueError(f"Target word is missing from CounterBalance.xlsx: {row['target_word']}")
        counterbalance_word_number = counterbalance_word_numbers[target_word]
        source_word_number = int(row["word_number"])

        manifest_rows.append(
            {
                "audio_file": audio_file,
                "audio_url": audio_url,
                "target_word": row["target_word"],
                "participant_id": row["proposed_speaker_id"],
                "l1_condition": l1,
                "pronunciation_condition": pronunciation,
                # Blank means the material can satisfy any A-J list slot with
                # the same L1, word_number, and pronunciation condition.
                "stimulus_list": "",
                "word_number": str(counterbalance_word_number),
                "condition": "main",
                "talker": global_speaker_id,
                "take_number": str(int(row["take_number"])),
                "spoken_form": row["target_word"],
                "practice_note": "",
                "source_format": f"osf_crosswalk_{path_mode}",
                "counterbalance_word_number": str(counterbalance_word_number),
                "source_word_number": str(source_word_number),
                "global_speaker_id": global_speaker_id,
                "old_speaker_id": row["old_speaker_id"],
                "proposed_speaker_id": row["proposed_speaker_id"],
                "old_relative_path": current_audio_file,
                "osf_audio_file": osf_audio_file,
                "original_pronunciation_label": row["original_pronunciation_label"],
                "source_pass_number": row["source_pass_number"],
                "trial_number": str(int(row["trial_number"])),
                "talker_label": row["talker_label"],
            }
        )

    return sorted(
        manifest_rows,
        key=lambda item: (
            item["l1_condition"],
            item["pronunciation_condition"],
            item["participant_id"],
            int(item["word_number"]),
            int(item["source_word_number"]),
            int(item["take_number"]),
            int(item["trial_number"]),
        ),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--file-crosswalk",
        type=Path,
        default=Path(__file__).resolve().parents[2]
        / "Stimuli"
        / "osf_stimuli_file_rename_crosswalk_20260703.csv",
    )
    parser.add_argument(
        "--speaker-crosswalk",
        type=Path,
        default=Path(__file__).resolve().parents[2]
        / "Stimuli"
        / "osf_speaker_id_crosswalk_20260703.csv",
    )
    parser.add_argument(
        "--counterbalance-workbook",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "stimuli" / "CounterBalance.xlsx",
        help="Workbook whose Sheet1 maps target words to counterbalance word numbers.",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parents[2]
        / "Stimuli"
        / "remote_manifest_production_osf_20260703.csv",
    )
    parser.add_argument(
        "--path-mode",
        choices=["osf", "current"],
        default="osf",
        help="Use OSF-standardized paths or current source paths in audio_file.",
    )
    parser.add_argument(
        "--audio-base-url",
        default="",
        help="Optional URL prefix for audio_url. Leave blank to use audio_file.",
    )
    args = parser.parse_args()

    file_rows = read_csv(args.file_crosswalk.expanduser().resolve())
    speaker_rows = read_csv(args.speaker_crosswalk.expanduser().resolve())
    counterbalance_word_numbers = read_counterbalance_word_numbers(
        args.counterbalance_workbook.expanduser().resolve(),
    )
    rows = build_manifest_rows(
        file_rows,
        speaker_lookup(speaker_rows),
        counterbalance_word_numbers,
        args.path_mode,
        args.audio_base_url.strip(),
    )
    out = args.out.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    write_csv(out, rows)

    print(f"manifest: {out}")
    print(f"rows: {len(rows)}")
    by_condition: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (row["l1_condition"], row["pronunciation_condition"])
        by_condition[key] = by_condition.get(key, 0) + 1
    for key, value in sorted(by_condition.items()):
        print(f"{key[0]}/{key[1]}: {value}")
    print("word_number: CounterBalance.xlsx lexical item number")
    print("source_word_number: word number parsed from the source audio filename")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
